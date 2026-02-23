#!/usr/bin/env python3
"""
Liest alle CSV-Dateien im Script-Verzeichnis (Semikolon-getrennt, mit
Multi-Line-Feldern) und erzeugt je eine XLSX-Datei mit:
  - Anrede, Name, Vorname, Geburtsdatum
  - Je einer Spalte pro Frage/Note aus "BewerbungenZusatzfragenBeantworteteFragenPflicht"
"""

import argparse
import csv
import glob
import os
import subprocess
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

GRADES_COLUMN = "BewerbungenZusatzfragenBeantworteteFragenPflicht"
GROUP_COLUMN = "Schüler:in Bildungsangebot Vollqualifizierter Schlüssel"

BASE_FIELDS = [
    ("Schüler:in Anrede Bezeichnung", "Anrede"),
    ("Schüler:in Name", "Name"),
    ("Schüler:in Vorname", "Vorname"),
    ("Schüler:in Geburtsdatum", "Geburtsdatum"),
    ("Schüler:in Sonderpädagogischer Förderbedarf", "Sonderpäd. Förderbedarf"),
    ("Schüler:in Bewerbung Prioritaetsrang", "Rang"),
]

# Extra field inserted at position 4 (5th column) for FG groups
FG_EXTRA_FIELD = ("Schüler:in abgebende Schule Schulgliederung", "Schulgliederung")

FIELD_TRANSFORMS = {
    "Schüler:in Sonderpädagogischer Förderbedarf": lambda v: "X" if v == "J" else "",
}

# Sprachliche Bewertungen (AV/SV) auf Zahlenwerte abbilden
BEWERTUNG_KEY = "Bitte geben Sie die Bewertung vom Zeugnis ein."
BEWERTUNG_RENAME = {1: "AV", 2: "SV"}
BEWERTUNG_WERTE = {
    "verdient besondere Anerkennung": 10,
    "entspricht den Erwartungen in vollem Umfang": 20,
    "entspricht den Erwartungen": 30,
    "entspricht den Erwartungen mit Einschränkungen": 40,
    "entspricht nicht den Erwartungen": 50,
}


def parse_answers(field_content):
    """Zerlegt ein mehrzeiliges Feld 'Schlüssel: Wert' in eine Liste von Tupeln.

    Doppelte Schlüssel im selben Datensatz werden durchnummeriert:
      Kunst: 3            -> ("Kunst", "3")
      Bewertung: gut      -> ("Bewertung", "gut")
      Bewertung: sehr gut -> ("Bewertung (2)", "sehr gut")
    """
    if not field_content or not field_content.strip():
        return []

    results = []
    seen = {}
    for line in field_content.split("\n"):
        line = line.strip()
        if not line or ": " not in line:
            continue
        key, value = line.split(": ", 1)
        key = key.strip()
        value = value.strip()

        count = seen.get(key, 0) + 1
        seen[key] = count

        # "Bitte geben Sie die Bewertung..." → AV / SV + Zahlenwert
        if key == BEWERTUNG_KEY:
            col_name = BEWERTUNG_RENAME.get(count, f"Bewertung ({count})")
            value = BEWERTUNG_WERTE.get(value, value)
        else:
            col_name = key if count == 1 else f"{key} ({count})"

        results.append((col_name, value))

    return results


def read_csv(csv_path):
    """Liest eine Semikolon-CSV mit quoted Multi-Line-Feldern."""
    encodings = ["utf-8-sig", "utf-8", "latin-1"]
    for enc in encodings:
        try:
            with open(csv_path, "r", encoding=enc) as f:
                reader = csv.DictReader(f, delimiter=";", quotechar='"')
                records = list(reader)
            return records
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise RuntimeError(f"Konnte {csv_path} mit keinem Encoding lesen.")


def write_xlsx(records, xlsx_path, is_fg=False):
    """Schreibt eine Liste von Datensätzen als XLSX."""
    # Basis-Felder ggf. um Schulgliederung erweitern
    base_fields = list(BASE_FIELDS)
    if is_fg:
        base_fields.insert(4, FG_EXTRA_FIELD)

    # Noten-Spalten aus allen Datensätzen dieser Gruppe sammeln
    all_answer_cols = []
    for rec in records:
        answers = parse_answers(rec.get(GRADES_COLUMN, ""))
        for col_name, _ in answers:
            if col_name not in all_answer_cols:
                all_answer_cols.append(col_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Bewerber"

    # Header
    header_labels = [label for _, label in base_fields] + all_answer_cols
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin"))

    for col_idx, label in enumerate(header_labels, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Datenzeilen
    for row_idx, rec in enumerate(records, start=2):
        for col_idx, (csv_col, _) in enumerate(base_fields, start=1):
            value = rec.get(csv_col, "")
            transform = FIELD_TRANSFORMS.get(csv_col)
            if transform:
                value = transform(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

        answers = parse_answers(rec.get(GRADES_COLUMN, ""))
        answer_dict = dict(answers)

        for col_idx, col_name in enumerate(all_answer_cols, start=len(base_fields) + 1):
            value = answer_dict.get(col_name, "")
            try:
                value = int(value)
            except (ValueError, TypeError):
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Spaltenbreiten
    for col_idx, label in enumerate(header_labels, start=1):
        width = max(len(str(label)), 8) + 2
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.auto_filter.ref = ws.dimensions
    wb.save(xlsx_path)
    return len(all_answer_cols)


def escape_latex(text):
    """Sonderzeichen für LaTeX escapen."""
    if not text:
        return ""
    text = str(text)
    text = text.replace("\\", r"\textbackslash{}")
    for char, repl in [
        ("&", r"\&"), ("%", r"\%"), ("$", r"\$"), ("#", r"\#"),
        ("_", r"\_"), ("{", r"\{"), ("}", r"\}"),
        ("~", r"\textasciitilde{}"), ("^", r"\textasciicircum{}"),
    ]:
        text = text.replace(char, repl)
    return text


def write_latex_pdf(records, pdf_path, bildungsgang, verbose=False, is_fg=False):
    """Erzeugt eine LaTeX-Übersichtstabelle als PDF (A4 Querformat)."""

    def g(rec, field):
        return escape_latex(rec.get(field, "").strip())

    # Sortierung: Rang (numerisch), dann Name, Vorname
    records = sorted(records, key=lambda r: (
        int(r.get("Schüler:in Bewerbung Prioritaetsrang", "") or 999),
        r.get("Schüler:in Name", "").lower(),
        r.get("Schüler:in Vorname", "").lower(),
    ))

    rows = []
    for rec in records:
        # Spalte 1: Name/Adresse
        name = f"{g(rec, 'Schüler:in Name')}, {g(rec, 'Schüler:in Vorname')}"
        gebdat = g(rec, "Schüler:in Geburtsdatum")
        strasse_nr = f"{g(rec, 'Schüler:in Straße')} {g(rec, 'Schüler:in Hausnummer')}".strip()
        plz_ort = f"{g(rec, 'Schüler:in Postleitzahl')} {g(rec, 'Schüler:in Wohnort')}".strip()
        ortsteil = g(rec, "Schüler:in Ortsteil")
        if ortsteil:
            plz_ort += f" {ortsteil}"
        col1 = r" \newline ".join(p for p in [name, gebdat, strasse_nr, plz_ort] if p)

        # Spalte 2: Kontakt
        col2 = r" \newline ".join(p for p in [
            g(rec, "Schüler:in Telefonnummer Hauptnummer"),
            g(rec, "Schüler:in Telefonnummer (weitere)"),
            g(rec, "Schüler:in E-Mail-Adresse"),
        ] if p)

        # Spalte 3: Qualifikation / Status
        hoechst = rec.get(
            "Schüler:in Qualifikation höchster Schulabschluss Kürzel", ""
        ).strip()
        letzt = rec.get(
            "Schüler:in Qualifikation letzter Schulabschluss Kürzel", ""
        ).strip()
        quali = escape_latex(hoechst if hoechst else letzt)

        rang = g(rec, "Schüler:in Bewerbung Prioritaetsrang")
        unterlagen = g(rec, "Schüler:in Bewerbung Unterlagen vollständig eingereicht")
        foerder = rec.get(
            "Schüler:in Sonderpädagogischer Förderbedarf", ""
        ).strip()

        col3_parts = []
        if quali:
            col3_parts.append(quali)
        col3_parts.append("~")            # Leerzeile als Trenner
        if rang:
            col3_parts.append(rang)
        if unterlagen:
            col3_parts.append(unterlagen)
        if foerder == "J":
            col3_parts.append(r"Förd.\,X")
        col3 = r" \newline ".join(col3_parts)

        if is_fg:
            col_sg = g(rec, "Schüler:in abgebende Schule Schulgliederung")
            rows.append(f"    {col1} & {col2} & {col3} & {col_sg} & & \\\\\n    \\hline")
        else:
            rows.append(f"    {col1} & {col2} & {col3} & & \\\\\n    \\hline")

    bg_esc = escape_latex(bildungsgang)
    bg_bez = escape_latex(records[0].get("Schüler:in Bildungsgang Bezeichnung", "").strip())
    rows_tex = "\n".join(rows)

    if is_fg:
        col_spec = r"|L{5.5cm}|L{4.5cm}|L{2cm}|L{1.5cm}|L{2cm}|R|"
        header_row = (
            r"\textbf{Name/Adresse} & \textbf{Kontakt} & "
            r"\textbf{\footnotesize Abschl. \newline Rang \newline vollst.Unterl.} & "
            r"\textbf{\footnotesize Schulgl.} & "
            r"\textbf{\footnotesize Zusage/ \newline Absage/ \newline Warteliste} & "
            r"\textbf{Bemerkungen} \\"
        )
    else:
        col_spec = r"|L{5.5cm}|L{4.5cm}|L{2cm}|L{2cm}|R|"
        header_row = (
            r"\textbf{Name/Adresse} & \textbf{Kontakt} & "
            r"\textbf{\footnotesize Abschl. \newline Rang \newline vollst.Unterl.} & "
            r"\textbf{\footnotesize Zusage/ \newline Absage/ \newline Warteliste} & "
            r"\textbf{Bemerkungen} \\"
        )

    tex = rf"""\documentclass[a4paper,landscape,10pt]{{article}}
\usepackage[left=1.5cm,right=1.5cm,top=2cm,bottom=1.5cm]{{geometry}}
\usepackage[utf8]{{inputenc}}
\usepackage[T1]{{fontenc}}
\usepackage[default]{{sourcesanspro}}
\usepackage[ngerman]{{babel}}
\usepackage{{xltabular}}
\usepackage{{array}}
\usepackage{{fancyhdr}}

\pagestyle{{fancy}}
\fancyhf{{}}
\fancyhead[C]{{\large\bfseries {bg_esc} -- {bg_bez}}}
\fancyfoot[C]{{\thepage}}
\renewcommand{{\headrulewidth}}{{0.4pt}}
\setlength{{\parindent}}{{0pt}}
\renewcommand{{\arraystretch}}{{1.2}}

\begin{{document}}

\newcolumntype{{L}}[1]{{>{{\raggedright\arraybackslash}}p{{#1}}}}
\newcolumntype{{R}}{{>{{\raggedright\arraybackslash}}X}}

\begin{{xltabular}}{{\textwidth}}{{{col_spec}}}
\hline
{header_row}
\hline
\endfirsthead
\hline
{header_row}
\hline
\endhead
\hline
\endfoot
{rows_tex}
\end{{xltabular}}

\end{{document}}
"""

    tex_path = pdf_path.replace(".pdf", ".tex")
    with open(tex_path, "w", encoding="utf-8") as f:
        f.write(tex)

    return compile_latex(tex_path, verbose)


def compile_latex(tex_path, verbose=False):
    """Kompiliert eine .tex-Datei zu PDF. Gibt True zurück bei Erfolg."""
    out_dir = os.path.dirname(tex_path) or "."
    base = os.path.splitext(tex_path)[0]
    pdf_path = base + ".pdf"
    log_path = base + ".log"

    try:
        with open(log_path, "w", encoding="utf-8") as logf:
            for _ in range(2):
                result = subprocess.run(
                    ["pdflatex", "-interaction=nonstopmode",
                     "-output-directory", out_dir, tex_path],
                    capture_output=True, text=True, timeout=30,
                )
                logf.write(result.stdout)
                if result.stderr:
                    logf.write(result.stderr)
    except FileNotFoundError:
        print("    WARNUNG: pdflatex nicht gefunden – nur .tex erzeugt.")
        return False

    # Hilfsdateien aufräumen
    for ext in (".aux", ".out"):
        aux = base + ext
        if os.path.exists(aux):
            os.remove(aux)

    if not verbose:
        for ext in (".tex", ".log"):
            path = base + ext
            if os.path.exists(path):
                os.remove(path)

    return os.path.exists(pdf_path)


def write_summary_pdf(stats, pdf_path, verbose=False):
    """Erzeugt ein Übersichts-PDF mit Bildungsgang und Anzahl Datensätze."""
    total = sum(count for _, count in stats)

    rows_tex = "\n".join(
        f"    {escape_latex(bg)} & {count} \\\\\n    \\hline"
        for bg, count in stats
    )

    tex = rf"""\documentclass[a4paper,10pt]{{article}}
\usepackage[left=2cm,right=2cm,top=2cm,bottom=2cm]{{geometry}}
\usepackage[utf8]{{inputenc}}
\usepackage[T1]{{fontenc}}
\usepackage[default]{{sourcesanspro}}
\usepackage[ngerman]{{babel}}
\usepackage{{array}}
\usepackage{{booktabs}}

\setlength{{\parindent}}{{0pt}}

\begin{{document}}

\begin{{center}}
{{\large\bfseries Übersicht exportierte Datensätze}}
\end{{center}}

\vspace{{1cm}}

\begin{{tabular}}{{|l|r|}}
\hline
\textbf{{Bildungsgang}} & \textbf{{Anzahl}} \\
\hline
{rows_tex}
\hline
\textbf{{Gesamt}} & \textbf{{{total}}} \\
\hline
\end{{tabular}}

\end{{document}}
"""

    tex_path = pdf_path.replace(".pdf", ".tex")
    with open(tex_path, "w", encoding="utf-8") as f:
        f.write(tex)

    return compile_latex(tex_path, verbose)


def process_csv(csv_path, verbose=False):
    """Verarbeitet eine CSV: gruppiert nach Bildungsgang, sortiert nach Name."""
    records = read_csv(csv_path)
    if not records:
        print(f"  Keine Datensätze in {csv_path} — übersprungen.")
        return []

    if GRADES_COLUMN not in records[0]:
        print(f"  Spalte '{GRADES_COLUMN}' nicht gefunden in {csv_path} — übersprungen.")
        return []

    # Nach Bildungsgang gruppieren
    groups = defaultdict(list)
    for rec in records:
        group_key = rec.get(GROUP_COLUMN, "").strip() or "unbekannt"
        groups[group_key].append(rec)

    out_dir = os.path.dirname(csv_path)
    stats = []

    for bildungsgang, group_records in sorted(groups.items()):
        # Nach Name, Vorname sortieren
        group_records.sort(key=lambda r: (
            r.get("Schüler:in Name", "").lower(),
            r.get("Schüler:in Vorname", "").lower(),
        ))

        is_fg = bildungsgang.startswith("FG")

        # Dateiname: Slash → Dash
        safe_name = bildungsgang.replace("/", "-")
        xlsx_path = os.path.join(out_dir, f"{safe_name}.xlsx")
        n_cols = write_xlsx(group_records, xlsx_path, is_fg=is_fg)
        print(f"  -> {safe_name}.xlsx  ({len(group_records)} Bewerber, {n_cols} Noten-Spalten)")

        # LaTeX-PDF-Übersicht
        pdf_path = os.path.join(out_dir, f"{safe_name}.pdf")
        if write_latex_pdf(group_records, pdf_path, bildungsgang, verbose, is_fg=is_fg):
            print(f"  -> {safe_name}.pdf")

        stats.append((bildungsgang, len(group_records)))

    print(f"  Gesamt: {len(records)} Datensätze in {len(stats)} Dateien")
    return stats


def main():
    parser = argparse.ArgumentParser(description="CSV zu XLSX + PDF konvertieren")
    parser.add_argument("-v", "--verbose", action="store_true",
                        help="LaTeX-Logfiles und .tex-Dateien behalten")
    args = parser.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    csv_files = sorted(glob.glob(os.path.join(script_dir, "*.csv")))

    if not csv_files:
        print("Keine CSV-Dateien gefunden.")
        return

    all_stats = []
    print(f"{len(csv_files)} CSV-Datei(en) gefunden:\n")
    for csv_path in csv_files:
        print(f"Verarbeite: {os.path.basename(csv_path)}")
        all_stats.extend(process_csv(csv_path, args.verbose))

    if all_stats:
        summary_path = os.path.join(script_dir, "uebersicht.pdf")
        if write_summary_pdf(all_stats, summary_path, args.verbose):
            print(f"\n  -> uebersicht.pdf  ({len(all_stats)} Bildungsgänge)")

    print("\nFertig.")


if __name__ == "__main__":
    main()
